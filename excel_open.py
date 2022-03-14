import pandas as pd
import openpyxl
import sys
# import excel_cell_to_list

from pandas.core.indexes.base import Index
sys.argv

class excel_file_open():
    def __init__(self):
        self.file_name1 = 'D:/UserArea/J0134661/Desktop/便利機能/excel_check/コピーFMEA試験正解表_次世代PDRV_OBD対象⊿8_masuo.xlsx'
        self.file_sheet_name1 = input('エクセルファイルのsheet名を記入>>')
        print(self.file_sheet_name1)

    def read_excel_file_pandas(self):
        excel1 = pd.read_excel(self.file_name1,sheet_name=self.file_sheet_name1)
        excel_file_name1_split = self.file_name1.split('.')
        excel_file_name1_split[0] = excel_file_name1_split[0].replace(excel_file_name1_split[0],excel_file_name1_split[0]+'_backup')
        excel_backup_file = '.'.join(excel_file_name1_split)
        excel1.to_excel(excel_backup_file)
        
        return excel1

    def excel_color(self,number,pyxl_excel):
        
        excel2_sheet = pyxl_excel[self.file_sheet_name1]
        fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
        excel2_sheet.cell(row=number+1, column=28).fill = fill
        
        return excel2_sheet

    def pyxl_excel_color(self):
        excel2 = openpyxl.load_workbook(self.file_name1)
        return excel2

    def close_pyxl_excel_color(self,pyxl_excel_color):
        pyxl_excel_color.save("D:/UserArea/J0134661/Desktop/便利機能/excel_check/result.xlsx")

    def get_cell(self,x_read,list_cell_excelx):
        for i in range(len(x_read)):
            list_cell_excelx.append(x_read.iat[i,0])

        return list_cell_excelx

class excel_only_open():
    def __init__(self):
        self.file_name1 = 'D:/UserArea/J0134661/Desktop/便利機能/excel_check/コピーFMEA試験正解表_次世代PDRV_OBD対象⊿8_masuo.xlsx'
        self.file_sheet_name1 = input('エクセルファイルのsheet名を記入>>')
        print(self.file_sheet_name1)

    def open_excel_file(self):
        self.excel_pyxl_open = openpyxl.load_workbook(self.file_name1)
        return self.excel_pyxl_open

    def open_excel_sheet(self,excel_pyxl_open):
        self.execel_pyxl_sheet = self.execel_pyxl_open[self.file_sheet_name1]
        return 

