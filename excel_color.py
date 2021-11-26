import openpyxl


## openpyxl用のファイル管理
class excel_fill_color():
    def __init__(self):
        pass

    def excel_color(self,number,pyxl_excel,sheet,col_num):
        excel2_sheet = pyxl_excel[sheet]
        fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='FF0000', bgColor='FF0000')
        excel2_sheet.cell(row=number+2, column=col_num+1).fill = fill
        
        return excel2_sheet

    def open_excel_color(self,file):
        excel2 = openpyxl.load_workbook(file)
        return excel2

    def close_pyxl_excel_color(self,pyxl_excel_color):
        pyxl_excel_color.save("D:/UserArea/J0134661/Desktop/便利機能/excel_check/excel_check/result.xlsx")
